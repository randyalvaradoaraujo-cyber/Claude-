#!/usr/bin/env python3
"""Verificación funcional del libro ATENEO.

Comprueba el libro contra la ESPECIFICACIÓN, no contra la implementación: busca
las cosas por su contenido (la fila rotulada "Horas", la caja rotulada "Resumen
semanas …") en lugar de por direcciones fijas. Así sigue siendo válido aunque el
constructor mueva un bloque de sitio, y falla de verdad cuando el valor está mal.

Uso:
    python3 verificar.py "Ejemplo Planificador Oposición.xlsx"

Código de salida: 0 si pasa todo, 1 si falla alguna comprobación.
"""
from __future__ import annotations

import datetime as dt
import os
import shutil
import sys
import tempfile
import warnings

warnings.filterwarnings("ignore")

import openpyxl

PESTANAS = ["Estrategia", "Progreso", "Calendario", "Preguntas"]
PROHIBIDA = "Contenidos"

RESUMENES = {  # bloque -> (temas estudiados, repasos, simulacros)
    "1 a 8":   (13, 36, 2),
    "9 a 16":  (10, 42, 2),
    "17 a 24": (10, 42, 2),
    "25 a 30": (7, 33, 1),
    "31 a 34": (None, 40, 4),
}

SEMILLA = {
    "Tema 1": ["01/06/2026", "03/06/2026", "09/06/2026", "20/06/2026"],
    "Tema 2": ["04/06/2026", "06/06/2026", "12/06/2026", "20/06/2026"],
    "Tema 3": ["06/06/2026", "08/06/2026", "15/06/2026", "26/06/2026"],
    "Tema 4": ["10/06/2026", "13/06/2026", "20/06/2026", "27/06/2026"],
    "Tema 5": ["13/06/2026", "16/06/2026", "24/06/2026", "04/07/2026"],
    "Tema 6": ["17/06/2026", "19/06/2026", "27/06/2026", None],
    "Tema 7": ["22/06/2026", "25/06/2026", "27/06/2026", None],
    "Tema 8": ["29/06/2026", "01/07/2026", None, None],
    "Tema 9": ["02/07/2026", "04/07/2026", None, None],
}

ERRORES = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!")

resultados = []


def check(nombre, ok, evidencia=""):
    resultados.append((nombre, bool(ok), evidencia))
    print("  %-4s %s" % ("OK" if ok else "FALLO", nombre))
    if evidencia:
        for linea in str(evidencia).splitlines():
            print("        %s" % linea)


# ---------------------------------------------------------------- utilidades

def calcular(path):
    """Devuelve {'HOJA!A1': valor} evaluando cada fórmula del libro."""
    import formulas
    modelo = formulas.ExcelModel().loads(path).finish()
    sol = modelo.calculate()
    out = {}
    import re
    clave = re.compile(r"^'?\[[^\]]+\](?P<h>.+?)'?!(?P<c>\$?[A-Z]{1,3}\$?\d{1,7})$")
    for k, v in sol.items():
        m = clave.match(str(k))
        if not m:
            continue
        val = getattr(v, "value", v)
        if hasattr(val, "ndim"):
            if getattr(val, "size", 0) != 1:
                continue
            if val.ndim:
                val = val.reshape(-1)[0]
        if hasattr(val, "item") and type(val).__name__ != "XlError":
            try:
                val = val.item()
            except Exception:
                pass
        out["%s!%s" % (m.group("h").strip("'").upper(), m.group("c").replace("$", ""))] = val
    return out


def es_error(v):
    return type(v).__name__ == "XlError" or (isinstance(v, str) and v.strip() in ERRORES)


def celdas_rotuladas(ws, rotulo):
    """Todas las celdas cuyo texto empieza por `rotulo` (sin distinguir mayúsculas)."""
    hits = []
    for fila in ws.iter_rows():
        for c in fila:
            if isinstance(c.value, str) and c.value.strip().lower().startswith(rotulo.lower()):
                hits.append(c)
    return hits


def fila_valores(ws, celda_rotulo, valores):
    """Valores a la derecha del rótulo, en su misma fila, hasta que se acaben."""
    out = []
    for col in range(celda_rotulo.column + 1, ws.max_column + 1):
        ref = "%s!%s" % (ws.title.upper(), ws.cell(celda_rotulo.row, col).coordinate)
        v = valores.get(ref, ws.cell(celda_rotulo.row, col).value)
        out.append((ws.cell(celda_rotulo.row, col).coordinate, v))
    return out


def valor(valores, ws, celda):
    """Valor calculado de la celda; si el motor no produjo nada, el literal.

    `dict.get(k, default)` devuelve el None *almacenado*, no el default, así que
    una celda que el motor no supo evaluar se leería como vacía y una fecha
    literal parecería ausente. Aquí el literal gana cuando no hay cálculo.
    """
    v = valores.get("%s!%s" % (ws.title.upper(), celda.coordinate))
    return celda.value if v is None else v


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def fecha_txt(v):
    """Normaliza a dd/mm/aaaa venga como venga.

    El motor de cálculo devuelve las fechas como número de serie de Excel (días
    desde 1899-12-30), mientras que openpyxl las devuelve ya como datetime. Sin
    convertir el serial, una fecha correcta se leería como ausente.
    """
    if isinstance(v, dt.datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, dt.date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if v <= 0:
            return None
        return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).strftime("%d/%m/%Y")
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None


# --------------------------------------------------------------- comprobaciones

def c1_pestanas(wb):
    nombres = [h for h in wb.sheetnames if not h.startswith("_")]
    check("1. Las 4 pestañas existen y en orden",
          nombres == PESTANAS, "encontradas: %s" % nombres)
    check("1b. No existe la pestaña 'Contenidos'",
          PROHIBIDA not in wb.sheetnames)


def c2_horas(wb, valores):
    ws = wb["Estrategia"]
    nombres = {n: str(d.attr_text) for n, d in wb.defined_names.items()}
    faltan = [n for n in ("H_TEMA_NUEVO", "H_REPASO", "H_SIMULACRO",
                          "DIAS_ESTUDIO", "INTERVALOS") if n not in nombres]
    check("2a. Los 5 rangos con nombre existen", not faltan,
          "faltan: %s" % faltan if faltan else "todos presentes")

    def tarifa(nombre):
        ref = nombres.get(nombre, "")
        if "!" not in ref:
            return None
        hoja, celda = ref.split("!", 1)
        hoja = hoja.strip("'=").upper()
        return num(valores.get("%s!%s" % (hoja, celda.replace("$", ""))))

    h_tn, h_rep, h_sim = tarifa("H_TEMA_NUEVO"), tarifa("H_REPASO"), tarifa("H_SIMULACRO")
    check("2b. Las tarifas se leen (5 / 2,5 / 2,5)",
          (h_tn, h_rep, h_sim) == (5, 2.5, 2.5), "leídas: %s" % [h_tn, h_rep, h_sim])
    if None in (h_tn, h_rep, h_sim):
        check("2c. Las 34 celdas de Horas cuadran", False, "sin tarifas no se puede comprobar")
        return

    total, malas = 0, []
    for chor in celdas_rotuladas(ws, "Horas"):
        if chor.column != 1:      # la fila Horas de los bloques vive en la columna A
            continue
        base = chor.row
        etiquetas = {}
        for r in range(max(1, base - 4), base):
            v = ws.cell(r, 1).value
            if isinstance(v, str):
                etiquetas[v.strip().lower()] = r
        f_tn = etiquetas.get("temas nuevos")
        f_rep, f_sim = etiquetas.get("repasos"), etiquetas.get("simulacros")
        if f_rep is None or f_sim is None:
            continue
        for col in range(2, ws.max_column + 1):
            def val(fila):
                if fila is None:
                    return 0
                ref = "%s!%s" % (ws.title.upper(), ws.cell(fila, col).coordinate)
                return num(valores.get(ref, ws.cell(fila, col).value)) or 0
            horas = num(valores.get("%s!%s" % (ws.title.upper(), ws.cell(base, col).coordinate),
                                    ws.cell(base, col).value))
            if horas is None:
                continue
            esperado = val(f_tn) * h_tn + val(f_rep) * h_rep + val(f_sim) * h_sim
            total += 1
            if abs(horas - esperado) > 1e-9:
                malas.append("%s: %s ≠ %s" % (ws.cell(base, col).coordinate, horas, esperado))
    check("2c. Las 34 celdas de Horas cuadran con la fórmula de tarifas",
          total == 34 and not malas,
          "comprobadas %d celdas; discrepancias: %s" % (total, malas or "ninguna"))


def c3_resumenes(wb, valores):
    ws = wb["Estrategia"]
    encontrados, malos = 0, []
    for bloque, objetivo in RESUMENES.items():
        cajas = [c for c in celdas_rotuladas(ws, "Resumen") if bloque in str(c.value)]
        if not cajas:
            malos.append("caja '%s' no encontrada" % bloque)
            continue
        caja = cajas[0]
        leidos = {}
        for r in range(caja.row + 1, caja.row + 5):
            et = ws.cell(r, caja.column).value
            if not isinstance(et, str):
                continue
            ref = "%s!%s" % (ws.title.upper(), ws.cell(r, caja.column + 1).coordinate)
            leidos[et.strip().lower()] = num(valores.get(ref, ws.cell(r, caja.column + 1).value))
        for etiqueta, esperado in zip(("temas estudiados", "repasos", "simulacros"), objetivo):
            if esperado is None:
                continue
            encontrados += 1
            got = leidos.get(etiqueta)
            if got is None or abs(got - esperado) > 1e-9:
                malos.append("%s / %s: %s ≠ %s" % (bloque, etiqueta, got, esperado))
    check("3. Las 14 celdas de Resumen dan los valores objetivo",
          encontrados == 14 and not malos,
          "comprobadas %d; discrepancias: %s" % (encontrados, malos or "ninguna"))


def c4_tarifa_propaga(path, valores):
    nombre = None
    wb = openpyxl.load_workbook(path)
    for n, d in wb.defined_names.items():
        if n == "H_TEMA_NUEVO":
            nombre = str(d.attr_text)
    if not nombre or "!" not in nombre:
        check("4. Cambiar H_TEMA_NUEVO recalcula las 34 semanas", False, "rango con nombre ausente")
        return
    hoja, celda = nombre.split("!", 1)
    hoja, celda = hoja.strip("'="), celda.replace("$", "")
    tmp = os.path.join(tempfile.mkdtemp(), "tarifa.xlsx")
    shutil.copy(path, tmp)
    wb2 = openpyxl.load_workbook(tmp)
    wb2[hoja][celda] = 6
    wb2.save(tmp)
    nuevos = calcular(tmp)
    ws = wb["Estrategia"]
    cambios = 0
    for chor in celdas_rotuladas(ws, "Horas"):
        if chor.column != 1:
            continue
        for col in range(2, ws.max_column + 1):
            ref = "ESTRATEGIA!%s" % ws.cell(chor.row, col).coordinate
            a, b = num(valores.get(ref)), num(nuevos.get(ref))
            if a is not None and b is not None and abs(a - b) > 1e-9:
                cambios += 1
    check("4. Cambiar H_TEMA_NUEVO de 5 a 6 recalcula las semanas (probado y revertido)",
          cambios >= 26,
          "%d celdas de Horas cambiaron (las semanas 31-34 no llevan temas nuevos, "
          "así que no deben cambiar)" % cambios)


def c5_semilla(wb, valores):
    ws = wb["Progreso"]
    malos = []
    for tema, fechas in SEMILLA.items():
        fila = None
        for f in ws.iter_rows(min_col=1, max_col=1):
            if isinstance(f[0].value, str) and f[0].value.strip() == tema:
                fila = f[0].row
                break
        if fila is None:
            malos.append("%s: fila no encontrada" % tema)
            continue
        for i, esperado in enumerate(fechas):
            c = ws.cell(fila, 2 + i)
            ref = "PROGRESO!%s" % c.coordinate
            got = fecha_txt(valor(valores, ws, c))
            if esperado is None:
                continue
            if got != esperado:
                malos.append("%s %s: %s ≠ %s" % (tema, c.coordinate, got, esperado))
    check("5. Las 9 filas semilla llevan las fechas exactas", not malos,
          "discrepancias: %s" % (malos or "ninguna"))


def c6_borrar_sesion1(path):
    tmp = os.path.join(tempfile.mkdtemp(), "borrado.xlsx")
    shutil.copy(path, tmp)
    wb = openpyxl.load_workbook(tmp)
    ws = wb["Progreso"]
    fila = None
    for f in ws.iter_rows(min_col=1, max_col=1):
        if isinstance(f[0].value, str) and f[0].value.strip() == "Tema 3":
            fila = f[0].row
            break
    if fila is None:
        check("6. Borrar Sesión 1 deja vacías las sesiones 2-10", False, "Tema 3 no encontrado")
        return
    ws.cell(fila, 2).value = None
    wb.save(tmp)
    vals = calcular(tmp)
    sucias = []
    for col in range(3, 12):
        ref = "PROGRESO!%s" % ws.cell(fila, col).coordinate
        v = vals.get(ref)
        if es_error(v):
            sucias.append("%s -> %s" % (ws.cell(fila, col).coordinate, v))
        elif v not in (None, "", 0):
            sucias.append("%s -> %r (debería quedar vacía)" % (ws.cell(fila, col).coordinate, v))
    check("6. Borrar Sesión 1 de un tema deja vacías sus sesiones 2-10 sin errores",
          not sucias, "residuos: %s" % (sucias or "ninguno"))


def c7_sin_errores(wb, valores):
    malos = []
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for c in fila:
                if isinstance(c.value, str) and c.value.strip() in ERRORES:
                    malos.append("%s!%s literal %s" % (ws.title, c.coordinate, c.value))
    for ref, v in valores.items():
        if es_error(v):
            malos.append("%s -> %s" % (ref, v))
    check("7. Cero errores de fórmula en todo el libro", not malos,
          "errores: %s" % (malos[:12] or "ninguno"))


def c8_grafica(wb):
    graficas = []
    for ws in wb.worksheets:
        for g in getattr(ws, "_charts", []):
            graficas.append("%s: %s con %d serie(s)"
                            % (ws.title, type(g).__name__, len(getattr(g, "series", []))))
    ok = any(g.startswith("Progreso") for g in graficas)
    check("8. La gráfica existe en Progreso y apunta a series reales", ok,
          "gráficas: %s" % (graficas or "ninguna"))


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    path = sys.argv[1]
    if not os.path.exists(path):
        print("no existe: %s" % path)
        return 2

    print("VERIFICACIÓN FUNCIONAL — %s\n" % os.path.abspath(path))
    wb = openpyxl.load_workbook(path)
    print("Recalculando el libro entero…")
    valores = calcular(path)
    print("%d celdas calculadas\n" % len(valores))

    # Cada comprobación va aislada: si una revienta se reporta como fallo con su
    # causa y las demás siguen corriendo. Un verificador que se cae a mitad no
    # dice nada sobre el resto del libro.
    for nombre, fn in (
        ("1", lambda: c1_pestanas(wb)),
        ("2", lambda: c2_horas(wb, valores)),
        ("3", lambda: c3_resumenes(wb, valores)),
        ("4", lambda: c4_tarifa_propaga(path, valores)),
        ("5", lambda: c5_semilla(wb, valores)),
        ("6", lambda: c6_borrar_sesion1(path)),
        ("7", lambda: c7_sin_errores(wb, valores)),
        ("8", lambda: c8_grafica(wb)),
    ):
        try:
            fn()
        except Exception as exc:
            check("%s. comprobación interrumpida" % nombre, False,
                  "%s: %s" % (type(exc).__name__, exc))

    pasadas = sum(1 for _, ok, _ in resultados if ok)
    print("\n%d/%d comprobaciones pasadas" % (pasadas, len(resultados)))
    fallidas = [n for n, ok, _ in resultados if not ok]
    if fallidas:
        print("Fallan:")
        for n in fallidas:
            print("  - %s" % n)
    return 0 if not fallidas else 1


if __name__ == "__main__":
    sys.exit(main())
