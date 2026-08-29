#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Proyecto ATENEO — constructor del libro «Ejemplo Planificador Oposición.xlsx».

Reconstruye con fidelidad forense el sistema de estudio de las 6 capturas de
`ref/`, siguiendo `ESPECIFICACION.md`.

Uso:
    python3 build/construir.py [ruta_de_salida.xlsx]

Es idempotente: reescribe el libro entero desde cero en cada ejecución, así que
ejecutarlo dos veces produce exactamente el mismo resultado.

Reglas que respeta (ver ESPECIFICACION.md):
  · exactamente 4 pestañas: Estrategia, Progreso, Calendario, Preguntas
  · toda cifra derivada es fórmula viva (ningún total escrito a mano)
  · ninguna constante de negocio dentro de una fórmula: tarifas, intervalos,
    días hábiles y ventanas viven en rangos con nombre
  · funciones prohibidas: XLOOKUP, FILTER, UNIQUE, SORT, SEQUENCE, XMATCH
  · tablas nativas de Excel (ListObject), fuente Arial 10, textos en español
"""
from __future__ import annotations

import datetime as dt
import os
import re
import sys
import zipfile

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.comments import Comment
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# CONFIG — paleta, tipografía y geometría.  Todo lo ajustable vive aquí.
# ---------------------------------------------------------------------------

FUENTE = "Arial"
TAM = 10

# Paleta: valores de ESPECIFICACION.md §Paleta (Anexo B del MEGA-PROMPT).
# Entre corchetes, el color medido a cuentagotas sobre el JPEG de `ref/`; las
# capturas provienen de una grabación con curva de contraste propia (el mismo
# verde mide 2E5D49 en ref/03-06 y 377158 en ref/01-02), así que no sirven como
# cuentagotas calibrado.  Ver SUPUESTOS.md §Paleta.
PAL = {
    "verde_cab":    "0B5C39",   # [medido 2E5D49 / 377158]
    "txt_cab":      "FFFFFF",
    "banda":        "F0F4F1",   # [medido F6F7F9]
    "escala_min":   "E6F2EA",   # [medido D3E7CE]
    "escala_max":   "7CC49A",   # [medido 88BC71]
    "fila_horas":   "F1F3F4",   # [medido F1F1F1]
    "res_cab":      "C8E6C9",   # [medido ACD19E]
    "res_cuerpo":   "E2F0E4",   # [medido D3E7CE]
    "rojo_cab":     "A61C00",   # captura manda: cabecera roja oscura con texto
                                # blanco.  ESPECIFICACION dice E6A9A0 (claro) —
                                # incompatible con el texto blanco de ref/06.
    "rojo_cuerpo":  "F6D9D5",   # [medido F3C5C7]
    "exito":        "38761D",
    "fallo":        "A61C00",
    "chip_ref_bg":  "F4C7C3", "chip_ref_tx": "A61C00",
    "chip_ana_bg":  "C9DAF8", "chip_ana_tx": "1C4587",
    "chip_evi_bg":  "FCE5CD", "chip_evi_tx": "B45F06",
    "pildora":      "F1F3F4",   # píldora gris del desplegable vacío (ref/01)
    "col_vacia":    "EFEFEF",   # columnas H/I sin semana del bloque 25-30
    "borde_rojo":   "A61C00",   # recuadros de agrupación (ronda 1)
    "borde_azul":   "1C4587",   # recuadros de agrupación (rondas 2-4)
    "borde_verde":  "38761D",   # recuadros de agrupación (ronda 5)
    "texto":        "000000",
    "gris_texto":   "5F6368",
}

ALTO_FILA_TABLA = 22.5      # 30 px — alto de fila de tabla nativa en la captura
ALTO_FILA_BASE = 15.75      # 21 px — alto por defecto fuera de las tablas
ANCHO_NORMAL = 13.57        # 100 px
ANCHO_ANCHO = 16.43         # 120 px

FMT_FECHA = "dd/mm/yyyy"

# ---------------------------------------------------------------------------
# Utilidades de formato
# ---------------------------------------------------------------------------


def f(color=None, bold=False, size=TAM, italic=False):
    return Font(name=FUENTE, size=size, bold=bold, italic=italic,
                color=color or PAL["texto"])


def relleno(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def lado(color, estilo="thin"):
    return Side(style=estilo, color=color)


def pinta(ws, coord, valor=None, fill=None, font=None, align=None, fmt=None):
    c = ws[coord]
    if valor is not None:
        c.value = valor
    if fill is not None:
        c.fill = relleno(fill)
    c.font = font if font is not None else f()
    if align is not None:
        c.alignment = align
    if fmt is not None:
        c.number_format = fmt
    return c


def anchos(ws, mapa):
    for letra, w in mapa.items():
        ws.column_dimensions[letra].width = w


def alto(ws, filas, h=ALTO_FILA_TABLA):
    for r in filas:
        ws.row_dimensions[r].height = h


def tabla(ws, nombre, ref, franjas=True):
    """Tabla nativa (ListObject).  El estilo de banda real lo ponen los
    rellenos explícitos; el ListObject aporta cabecera con filtro y la
    semántica de tabla que pide la especificación."""
    t = Table(displayName=nombre, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
        showRowStripes=franjas, showColumnStripes=False)
    ws.add_table(t)
    return t


def nombre_rango(wb, nombre, destino):
    wb.defined_names.add(DefinedName(nombre, attr_text=destino))


ALI_IZQ = Alignment(horizontal="left", vertical="center")
ALI_DER = Alignment(horizontal="right", vertical="center")
ALI_CEN = Alignment(horizontal="center", vertical="center")
ALI_ARR = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# DATOS LITERALES (Anexo A / capturas ref/03 y ref/06)
# ---------------------------------------------------------------------------

TARIFAS = [("Tema nuevo", 5, "H_TEMA_NUEVO"),
           ("Repaso activo", 2.5, "H_REPASO"),
           ("Simulacro escrito", 2.5, "H_SIMULACRO")]

# (título, fila de cabecera, nº de semanas, primera semana, rojo)
BLOQUES = [
    ("Semanas 1 a 8",   7,  8,  1,  False),
    ("Semanas 9 a 16",  13, 8,  9,  False),
    ("Semanas 17 a 24", 19, 8,  17, False),
    ("Semanas 25 a 30", 25, 6,  25, False),
    ("Semanas 31 a 34", 31, 4,  31, True),
]

DATOS_BLOQUE = {
    7:  {"Temas nuevos": [3, 2, 1, 2, 1, 1, 1, 2],
         "Repasos":      [2, 4, 6, 3, 6, 6, 6, 3],
         "Simulacros":   [0, 0, 0, 1, 0, 0, 0, 1]},
    13: {"Temas nuevos": [1, 1, 1, 2, 1, 1, 1, 2],
         "Repasos":      [6, 6, 6, 3, 6, 6, 6, 3],
         "Simulacros":   [0, 0, 0, 1, 0, 0, 0, 1]},
    19: {"Temas nuevos": [1, 1, 1, 2, 1, 1, 1, 2],
         "Repasos":      [6, 6, 6, 3, 6, 6, 6, 3],
         "Simulacros":   [0, 0, 0, 1, 0, 0, 0, 1]},
    25: {"Temas nuevos": [1, 1, 1, 2, 1, 1],
         "Repasos":      [6, 6, 6, 3, 6, 6],
         "Simulacros":   [0, 0, 0, 1, 0, 0]},
    31: {"Repasos":      [10, 10, 10, 10],
         "Simulacros":   [1, 1, 1, 1]},
}

# Las dos columnas sin semana del bloque 25-30 (ref/06): fondo gris y un "-".
# El nombre de columna de un ListObject no puede repetirse en Excel, así que la
# segunda usa un guión corto tipográficamente equivalente (ver SUPUESTOS.md).
GUION = "-"          # lo que se ve en el cuerpo de esas dos columnas
# El nombre de columna de un ListObject no puede repetirse en Excel, así que la
# cabecera de la segunda usa un guión corto tipográficamente equivalente y de
# color igual al fondo (ver SUPUESTOS.md).
CAB_VACIA_H, CAB_VACIA_I = "-", "–"


def construir_estrategia(wb):
    ws = wb.create_sheet("Estrategia")
    ws.sheet_format.defaultRowHeight = ALTO_FILA_BASE
    anchos(ws, {"A": ANCHO_NORMAL, "B": ANCHO_ANCHO, "C": ANCHO_NORMAL,
                "D": ANCHO_NORMAL, "E": ANCHO_NORMAL, "F": ANCHO_NORMAL,
                "G": ANCHO_NORMAL, "H": ANCHO_NORMAL, "I": ANCHO_NORMAL,
                "J": ANCHO_NORMAL, "K": ANCHO_ANCHO, "L": ANCHO_NORMAL})

    fcab = f(PAL["txt_cab"])

    # --- Tabla_1: tarifas ------------------------------------------------
    pinta(ws, "B2", "Columna 1", PAL["verde_cab"], fcab, ALI_IZQ)
    pinta(ws, "C2", "Horas", PAL["verde_cab"], fcab, ALI_IZQ)
    for i, (etiqueta, horas, _n) in enumerate(TARIFAS):
        r = 3 + i
        banda = PAL["banda"] if i % 2 else "FFFFFF"
        pinta(ws, "B%d" % r, etiqueta, banda, f(), ALI_IZQ)
        pinta(ws, "C%d" % r, horas, banda, f(), ALI_DER)
    tabla(ws, "Tabla_1", "B2:C5")
    alto(ws, range(2, 6))

    # --- Bloques semanales ------------------------------------------------
    for titulo, r0, nsem, sem1, rojo in BLOQUES:
        col_cab = PAL["rojo_cab"] if rojo else PAL["verde_cab"]
        ultima = 1 + nsem                      # índice de la última columna con semana
        if r0 == 25:
            ultima = 9                          # A..I: incluye las dos vacías
        filas = (["Repasos", "Simulacros", "Horas"] if rojo
                 else ["Temas nuevos", "Repasos", "Simulacros", "Horas"])

        # cabecera
        pinta(ws, "A%d" % r0, "Tarea", col_cab, fcab, ALI_IZQ)
        for k in range(nsem):
            pinta(ws, "%s%d" % (get_column_letter(2 + k), r0),
                  "Semana %d" % (sem1 + k), col_cab, fcab, ALI_IZQ)
        if r0 == 25:
            pinta(ws, "H25", CAB_VACIA_H, PAL["col_vacia"], f(PAL["col_vacia"]), ALI_IZQ)
            pinta(ws, "I25", CAB_VACIA_I, PAL["col_vacia"], f(PAL["col_vacia"]), ALI_IZQ)

        # cuerpo
        for i, etiqueta in enumerate(filas):
            r = r0 + 1 + i
            es_horas = (etiqueta == "Horas")
            banda = (PAL["fila_horas"] if es_horas
                     else (PAL["banda"] if i % 2 else "FFFFFF"))
            pinta(ws, "A%d" % r, etiqueta, banda, f(), ALI_IZQ)
            for k in range(nsem):
                col = get_column_letter(2 + k)
                celda = "%s%d" % (col, r)
                if es_horas:
                    if rojo:
                        val = "=%s%d*H_REPASO+%s%d*H_SIMULACRO" % (
                            col, r0 + 1, col, r0 + 2)
                    elif r0 == 25:
                        # Las columnas H/I no tienen semana y muestran "-": la
                        # misma fórmula guardada las respeta y mantiene toda la
                        # fila con un único patrón.
                        val = ('=IF(%s%d=%s,%s,%s%d*H_TEMA_NUEVO+%s%d*H_REPASO'
                               '+%s%d*H_SIMULACRO)' % (
                                   col, r0 + 1, '"-"', '"-"',
                                   col, r0 + 1, col, r0 + 2, col, r0 + 3))
                    else:
                        val = ("=%s%d*H_TEMA_NUEVO+%s%d*H_REPASO+%s%d*H_SIMULACRO"
                               % (col, r0 + 1, col, r0 + 2, col, r0 + 3))
                    pinta(ws, celda, val, banda, f(), ALI_DER)
                else:
                    pinta(ws, celda, DATOS_BLOQUE[r0][etiqueta][k], banda, f(), ALI_DER)
            if r0 == 25:
                for col in ("H", "I"):
                    if es_horas:
                        val = ('=IF(%s%d=%s,%s,%s%d*H_TEMA_NUEVO+%s%d*H_REPASO'
                               '+%s%d*H_SIMULACRO)' % (
                                   col, r0 + 1, '"-"', '"-"',
                                   col, r0 + 1, col, r0 + 2, col, r0 + 3))
                        pinta(ws, "%s%d" % (col, r), val, PAL["col_vacia"], f(), ALI_IZQ)
                    else:
                        pinta(ws, "%s%d" % (col, r), GUION, PAL["col_vacia"], f(), ALI_IZQ)

        ref = "A%d:%s%d" % (r0, get_column_letter(ultima), r0 + len(filas))
        tabla(ws, titulo.replace(" ", "_"), ref)
        alto(ws, range(r0, r0 + len(filas) + 1))

    # --- Escala de color de la fila «Temas nuevos» -------------------------
    # Un único rango múltiple: el mínimo y el máximo se calculan sobre la unión
    # de los cuatro bloques, así que no hace falta escribir umbrales.
    ws.conditional_formatting.add(
        "B8:I8 B14:I14 B20:I20 B26:G26",
        ColorScaleRule(start_type="min", start_color=PAL["escala_min"],
                       end_type="max", end_color=PAL["escala_max"]))

    # --- Cajas Resumen ----------------------------------------------------
    resumenes = [
        ("K", 7,  "Resumen semanas 1 a 8",
         [("Temas estudiados", "B8:I8"), ("Repasos", "B9:I9"), ("Simulacros", "B10:I10")], False),
        ("K", 13, "Resumen semanas 9 a 16",
         [("Temas estudiados", "B14:I14"), ("Repasos", "B15:I15"), ("Simulacros", "B16:I16")], False),
        ("K", 19, "Resumen semanas 17 a 24",
         [("Temas estudiados", "B20:I20"), ("Repasos", "B21:I21"), ("Simulacros", "B22:I22")], False),
        ("K", 25, "Resumen semanas 25 a 30",
         [("Temas estudiados", "B26:G26"), ("Repasos", "B27:G27"), ("Simulacros", "B28:G28")], False),
        ("G", 31, "Resumen semanas 31 a 34",
         [("Repasos", "B32:E32"), ("Simulacros", "B33:E33")], True),
    ]
    for col_izq, r0, titulo, filas, rojo in resumenes:
        c1 = col_izq
        c2 = get_column_letter(ws[c1 + "1"].column + 1)
        cab = PAL["rojo_cab"] if rojo else PAL["res_cab"]
        cuerpo = PAL["rojo_cuerpo"] if rojo else PAL["res_cuerpo"]
        fcabres = f(PAL["txt_cab"], bold=True) if rojo else f(bold=True)
        pinta(ws, "%s%d" % (c1, r0), titulo, cab, fcabres, ALI_IZQ)
        pinta(ws, "%s%d" % (c2, r0), None, cab, fcabres, ALI_IZQ)
        for i, (etiqueta, rango) in enumerate(filas):
            r = r0 + 1 + i
            pinta(ws, "%s%d" % (c1, r), etiqueta, cuerpo, f(), ALI_IZQ)
            pinta(ws, "%s%d" % (c2, r), "=SUM(%s)" % rango, cuerpo, f(), ALI_DER)

    return ws


# ---------------------------------------------------------------------------
# PESTAÑA 2 — Progreso
# ---------------------------------------------------------------------------

N_TEMAS = 40
N_SESIONES = 10

DIAS_ESTUDIO_LISTA = ["Lunes", "Martes", "Miércoles", "Jueves"]
DIAS_SEMANA_LISTA = ["Lunes", "Martes", "Miércoles", "Jueves",
                     "Viernes", "Sábado", "Domingo"]
INTERVALOS_LISTA = [2, 7, 14, 30, 60, 90, 120, 150, 180]

D = lambda dd, mm: dt.date(2026, mm, dd)

# Fechas semilla, §8.2.  Sesión 1 es entrada directa; las sesiones 2-10 van al
# bloque de semillas y sobrescriben la propuesta del motor.
SEMILLAS = {
    1: [D(1, 6),  D(3, 6),  D(9, 6),  D(20, 6)],
    2: [D(4, 6),  D(6, 6),  D(12, 6), D(20, 6)],
    3: [D(6, 6),  D(8, 6),  D(15, 6), D(26, 6)],
    4: [D(10, 6), D(13, 6), D(20, 6), D(27, 6)],
    5: [D(13, 6), D(16, 6), D(24, 6), D(4, 7)],
    6: [D(17, 6), D(19, 6), D(27, 6)],
    7: [D(22, 6), D(25, 6), D(27, 6)],
    8: [D(29, 6), D(1, 7)],
    9: [D(2, 7),  D(4, 7)],
}

# Recuadros de agrupación de ref/04 y ref/05: cada banda es una semana natural
# del calendario (lunes a domingo) y agrupa todas las celdas cuya fecha cae en
# ella.  Verificado celda a celda: las 30 fechas semilla quedan repartidas en
# estas cinco bandas sin sobras.  Coordenadas (fila de hoja, columna).
BANDAS = [
    ("borde_rojo",  [(2, "B"), (2, "C"), (3, "B"), (3, "C"), (4, "B")]),
    ("borde_azul",  [(2, "D"), (3, "D"), (4, "C"), (5, "B"), (5, "C"), (6, "B")]),
    ("borde_azul",  [(2, "E"), (3, "E"), (4, "D"), (5, "D"),
                     (6, "C"), (7, "C"), (7, "B")]),
    ("borde_azul",  [(4, "E"), (5, "E"), (6, "D"), (7, "D"), (8, "D"),
                     (8, "C"), (8, "B")]),
    ("borde_verde", [(6, "E"), (9, "B"), (9, "C"), (10, "B"), (10, "C")]),
]

# Columnas del motor invisible (ocultas).  ESPECIFICACION.md las situaba en una
# hoja `_Motor`, pero el libro debe tener exactamente 4 pestañas, así que viven
# aquí, fuera de la vista.  Ver SUPUESTOS.md.
COL_SEM = 18          # R: semillas de Sesión 2 (…Z para Sesión 10)
COL_ESP = 28          # AB: matriz espejo de estados (…AK)
COL_NUM = 52          # AZ: espejo numérico de fechas (…BI)
COL_AGG = 42          # AP: agregación semanal para la gráfica (…AX)
N_SEMANAS_GRAF = 34


def _rango_esp():
    return "$%s$2:$%s$%d" % (get_column_letter(COL_ESP),
                             get_column_letter(COL_ESP + N_SESIONES - 1), N_TEMAS + 1)


def _rango_num():
    return "$%s$2:$%s$%d" % (get_column_letter(COL_NUM),
                             get_column_letter(COL_NUM + N_SESIONES - 1), N_TEMAS + 1)


def construir_progreso(wb):
    ws = wb.create_sheet("Progreso")
    ws.sheet_format.defaultRowHeight = ALTO_FILA_BASE
    for i in range(1, 14):
        ws.column_dimensions[get_column_letter(i)].width = ANCHO_NORMAL
    fcab = f(PAL["txt_cab"])
    ultima = get_column_letter(1 + N_SESIONES)          # K

    # --- Tabla_2 ---------------------------------------------------------
    pinta(ws, "A1", "Temas", PAL["verde_cab"], fcab, ALI_IZQ)
    for s in range(1, N_SESIONES + 1):
        pinta(ws, "%s1" % get_column_letter(1 + s), "Sesión %d" % s,
              PAL["verde_cab"], fcab, ALI_IZQ)

    for t in range(1, N_TEMAS + 1):
        r = t + 1
        banda = PAL["banda"] if r % 2 else "FFFFFF"
        pinta(ws, "A%d" % r, "Tema %d" % t, banda, f(), ALI_IZQ)
        # Sesión 1: entrada manual del usuario
        v1 = SEMILLAS[t][0] if t in SEMILLAS else None
        pinta(ws, "B%d" % r, v1, banda, f(), ALI_DER, FMT_FECHA)
        # Sesiones 2-10: motor de repetición espaciada con semilla que manda
        for s in range(2, N_SESIONES + 1):
            col = get_column_letter(1 + s)
            prev = get_column_letter(s)
            sem = get_column_letter(COL_SEM + s - 2)
            paso = "%s%d+INDEX(INTERVALOS,COLUMN()-COLUMN($B$1))" % (prev, r)
            formula = ('=IF($B%d="","",IF(%s%d<>"",%s%d,'
                       '(%s)+INDEX(DESPLAZAMIENTO,WEEKDAY(%s,2))))'
                       % (r, sem, r, sem, r, paso, paso))
            pinta(ws, "%s%d" % (col, r), formula, banda, f(), ALI_DER, FMT_FECHA)

    ws["B1"].comment = Comment(
        "Sesión 1 es la única celda de entrada manual de cada tema: se escribe "
        "la fecha en que se estudió el tema por primera vez. Las sesiones 2 a 10 "
        "son fórmulas que proponen la fecha siguiente y pueden sobrescribirse. "
        "Por eso esta columna mezcla valores escritos con filas de fórmulas: es "
        "una excepción de diseño, no un valor pegado por error.",
        "Proyecto ATENEO", height=150, width=320)
    tabla(ws, "Tabla_2", "A1:%s%d" % (ultima, N_TEMAS + 1))
    alto(ws, range(1, N_TEMAS + 2))

    # --- Recuadros de agrupación (rondas de repaso) -----------------------
    aristas = {}
    for color, celdas in BANDAS:
        dentro = set(celdas)
        for (r, c) in celdas:
            ci = ws[c + "1"].column
            izq = (r, get_column_letter(ci - 1))
            der = (r, get_column_letter(ci + 1))
            if izq not in dentro:
                aristas.setdefault(("V", r, ci - 1), color)
            if der not in dentro:
                aristas.setdefault(("V", r, ci), color)
            if (r + 1, c) not in dentro:
                aristas.setdefault(("H", r, ci), color)
            # el borde superior sólo se dibuja entre filas de datos: en la
            # captura no hay línea de color bajo la cabecera verde
            if (r - 1, c) not in dentro and r - 1 >= 2:
                aristas.setdefault(("H", r - 1, ci), color)

    def _mete(celda, **kw):
        b = celda.border
        celda.border = Border(
            left=kw.get("left", b.left), right=kw.get("right", b.right),
            top=kw.get("top", b.top), bottom=kw.get("bottom", b.bottom))

    for (tipo, r, ci), color in aristas.items():
        s = lado(PAL[color])
        if tipo == "V":
            _mete(ws.cell(row=r, column=ci), right=s)
            _mete(ws.cell(row=r, column=ci + 1), left=s)
        else:
            _mete(ws.cell(row=r, column=ci), bottom=s)
            _mete(ws.cell(row=r + 1, column=ci), top=s)

    # --- Zona de parámetros (columna M, visible y editable) ---------------
    pinta(ws, "M1", "Días de estudio", None, f(bold=True), ALI_IZQ)
    for i, dia in enumerate(DIAS_ESTUDIO_LISTA):
        pinta(ws, "M%d" % (2 + i), dia, None, f(), ALI_IZQ)
    pinta(ws, "M8", "Intervalos (días)", None, f(bold=True), ALI_IZQ)
    for i, v in enumerate(INTERVALOS_LISTA):
        pinta(ws, "M%d" % (9 + i), v, None, f(), ALI_DER)

    # --- Motor invisible ---------------------------------------------------
    pinta(ws, "O1", "Día de la semana", None, f(bold=True), ALI_IZQ)
    pinta(ws, "P1", "Días hasta el siguiente día hábil", None, f(bold=True), ALI_IZQ)
    for i, dia in enumerate(DIAS_SEMANA_LISTA):
        pinta(ws, "O%d" % (2 + i), dia, None, f(), ALI_IZQ)
        # Cadena de IF que recorre los días siguientes hasta dar con uno hábil.
        # Se calcula desde DIAS_ESTUDIO: si el usuario cambia los días hábiles,
        # el desplazamiento se recalcula solo.  Sin constantes de negocio.
        expr = "0"
        for salto in range(6, -1, -1):
            ref = "$O$%d" % (2 + (i + salto) % 7)
            expr = "IF(COUNTIF(DIAS_ESTUDIO,%s)>0,%d,%s)" % (ref, salto, expr)
        pinta(ws, "P%d" % (2 + i), "=" + expr, None, f(), ALI_DER)

    col_sem = get_column_letter(COL_SEM)
    pinta(ws, "%s1" % col_sem,
          "Semillas de Sesión 2 a 10 (sobrescriben la propuesta del motor)",
          None, f(bold=True), ALI_IZQ)
    for t, fechas in SEMILLAS.items():
        for k, fecha in enumerate(fechas[1:]):
            pinta(ws, "%s%d" % (get_column_letter(COL_SEM + k), t + 1),
                  fecha, None, f(), ALI_DER, FMT_FECHA)

    col_esp = get_column_letter(COL_ESP)
    pinta(ws, "%s1" % col_esp,
          "Matriz espejo de estados por sesión (Éxito / Fallo / vacío)",
          None, f(bold=True), ALI_IZQ)

    col_num = get_column_letter(COL_NUM)
    pinta(ws, "%s1" % col_num,
          "Espejo numérico de fechas (0 = sin fecha)", None, f(bold=True), ALI_IZQ)
    for t in range(1, N_TEMAS + 1):
        r = t + 1
        for s in range(1, N_SESIONES + 1):
            origen = get_column_letter(1 + s)
            pinta(ws, "%s%d" % (get_column_letter(COL_NUM + s - 1), r),
                  '=IF(%s%d="",0,%s%d)' % (origen, r, origen, r), None, f(), ALI_DER)

    # etiquetas y parámetros con nombre del motor
    pinta(ws, "AM1", "Etiqueta de sesión superada", None, f(bold=True), ALI_IZQ)
    pinta(ws, "AN1", "Éxito", None, f(), ALI_IZQ)
    pinta(ws, "AM2", "Etiqueta de sesión fallida", None, f(bold=True), ALI_IZQ)
    pinta(ws, "AN2", "Fallo", None, f(), ALI_IZQ)
    pinta(ws, "AM3", "Lunes de la primera sesión", None, f(bold=True), ALI_IZQ)
    rejilla = "$B$2:$%s$%d" % (get_column_letter(1 + N_SESIONES), N_TEMAS + 1)
    # MIN sobre la rejilla real (ignora el texto vacío que devuelven las
    # fórmulas); el espejo numérico vale 0 y falsearía el mínimo.
    pinta(ws, "AN3",
          '=IF(MIN(%s)=0,"",MIN(%s)-WEEKDAY(MIN(%s),2)+1)'
          % (rejilla, rejilla, rejilla), None, f(), ALI_DER, FMT_FECHA)
    pinta(ws, "AM4", "Días por semana", None, f(bold=True), ALI_IZQ)
    pinta(ws, "AN4", 7, None, f(), ALI_DER)
    pinta(ws, "AM5", "Ventana «próximos días»", None, f(bold=True), ALI_IZQ)
    pinta(ws, "AN5", 7, None, f(), ALI_DER)

    # --- Agregación semanal que alimenta la gráfica ------------------------
    cab_agg = ["Semana", "Inicio", "Fin", "Éxitos", "Fallos",
               "Tasa de éxito", "Completadas", "Acumulado", "Media móvil 3"]
    for i, texto in enumerate(cab_agg):
        pinta(ws, "%s1" % get_column_letter(COL_AGG + i), texto, None,
              f(bold=True), ALI_IZQ)
    c = {n: get_column_letter(COL_AGG + i) for i, n in enumerate(
        ["sem", "ini", "fin", "exi", "fal", "tasa", "com", "acu", "mm3"])}
    for k in range(1, N_SEMANAS_GRAF + 1):
        r = k + 1
        pinta(ws, "%s%d" % (c["sem"], r), k, None, f(), ALI_DER)
        pinta(ws, "%s%d" % (c["ini"], r),
              '=IF($AN$3="","",$AN$3+(%s%d-1)*DIAS_POR_SEMANA)' % (c["sem"], r),
              None, f(), ALI_DER, FMT_FECHA)
        pinta(ws, "%s%d" % (c["fin"], r),
              '=IF(%s%d="","",%s%d+DIAS_POR_SEMANA-1)' % (c["ini"], r, c["ini"], r),
              None, f(), ALI_DER, FMT_FECHA)
        for clave, etiqueta in (("exi", "ESTADO_EXITO"), ("fal", "ESTADO_FALLO")):
            pinta(ws, "%s%d" % (c[clave], r),
                  '=IF(%s%d="",0,SUMPRODUCT((%s=%s)*(%s>=%s%d)*(%s<=%s%d)))'
                  % (c["ini"], r, _rango_esp(), etiqueta,
                     _rango_num(), c["ini"], r, _rango_num(), c["fin"], r),
                  None, f(), ALI_DER)
        # sin IFERROR: se trata el único fallo posible (semana sin sesiones)
        pinta(ws, "%s%d" % (c["tasa"], r),
              '=IF(%s%d+%s%d=0,"",%s%d/(%s%d+%s%d))'
              % (c["exi"], r, c["fal"], r, c["exi"], r, c["exi"], r, c["fal"], r),
              None, f(), ALI_DER, "0%")
        pinta(ws, "%s%d" % (c["com"], r),
              "=%s%d+%s%d" % (c["exi"], r, c["fal"], r), None, f(), ALI_DER)
        pinta(ws, "%s%d" % (c["acu"], r),
              "=SUM($%s$2:%s%d)" % (c["com"], c["com"], r), None, f(), ALI_DER)
        if k >= 3:
            pinta(ws, "%s%d" % (c["mm3"], r),
                  '=IF(COUNTIFS(%s%d:%s%d,">=0")=0,"",'
                  'SUM(%s%d:%s%d)/COUNTIFS(%s%d:%s%d,">=0"))'
                  % (c["tasa"], r - 2, c["tasa"], r,
                     c["tasa"], r - 2, c["tasa"], r,
                     c["tasa"], r - 2, c["tasa"], r), None, f(), ALI_DER, "0%")

    # --- Tarjetas KPI ------------------------------------------------------
    esp, num = _rango_esp(), _rango_num()
    tarjetas = [
        ("A", "Tasa de éxito global",
         '=IF(COUNTIF(%s,ESTADO_EXITO)+COUNTIF(%s,ESTADO_FALLO)=0,"",'
         'COUNTIF(%s,ESTADO_EXITO)/(COUNTIF(%s,ESTADO_EXITO)+COUNTIF(%s,ESTADO_FALLO)))'
         % (esp, esp, esp, esp, esp), "0%"),
        ("C", "Racha actual",
         '=SUMPRODUCT((%s=ESTADO_EXITO)*(%s>SUMPRODUCT(MAX((%s=ESTADO_FALLO)*%s))))'
         % (esp, num, esp, num), "0"),
        ("E", "Sesiones vencidas",
         '=SUMPRODUCT((%s>0)*(%s<HOY)*(%s=""))' % (num, num, esp), "0"),
    ]
    for col, titulo, formula, fmt in tarjetas:
        der = get_column_letter(ws[col + "1"].column + 1)
        pinta(ws, "%s43" % col, titulo, PAL["res_cab"], f(bold=True), ALI_IZQ)
        pinta(ws, "%s43" % der, None, PAL["res_cab"], f(), ALI_IZQ)
        pinta(ws, "%s44" % col, formula, PAL["res_cuerpo"],
              f(size=16, bold=True), ALI_CEN, fmt)
        pinta(ws, "%s44" % der, None, PAL["res_cuerpo"], f(), ALI_CEN)
    alto(ws, [43, 44])
    ws.conditional_formatting.add("A44", CellIsRule(
        operator="greaterThanOrEqual", formula=["0.8"],
        font=Font(name=FUENTE, size=16, bold=True, color=PAL["exito"])))
    ws.conditional_formatting.add("A44", CellIsRule(
        operator="between", formula=["0.6", "0.7999999"],
        font=Font(name=FUENTE, size=16, bold=True, color="B45F06")))
    ws.conditional_formatting.add("A44", CellIsRule(
        operator="lessThan", formula=["0.6"],
        font=Font(name=FUENTE, size=16, bold=True, color=PAL["fallo"])))
    ws["C44"].font = f(PAL["exito"], bold=True, size=16)
    ws.conditional_formatting.add("E44", CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(name=FUENTE, size=16, bold=True, color=PAL["fallo"])))

    # --- Gráfica -----------------------------------------------------------
    fila_ini, fila_fin = 1, N_SEMANAS_GRAF + 1
    col = {n: COL_AGG + i for i, n in enumerate(
        ["sem", "ini", "fin", "exi", "fal", "tasa", "com", "acu", "mm3"])}
    g1 = LineChart()
    g1.title = "📈 Curva de dominio — retención por semana"
    g1.style = 2
    g1.height, g1.width = 9.5, 22
    g1.y_axis.title = "Tasa de éxito"
    g1.x_axis.title = "Semana"
    g1.legend.position = "b"
    g1.x_axis.axPos = "b"
    for clave in ("tasa", "mm3"):
        g1.add_data(Reference(ws, min_col=col[clave], min_row=fila_ini,
                              max_row=fila_fin), titles_from_data=True)
    g1.set_categories(Reference(ws, min_col=col["sem"], min_row=fila_ini + 1,
                                max_row=fila_fin))
    g2 = LineChart()
    g2.add_data(Reference(ws, min_col=col["acu"], min_row=fila_ini,
                          max_row=fila_fin), titles_from_data=True)
    g2.y_axis.axId = 200
    g2.y_axis.title = "Sesiones acumuladas"
    g2.y_axis.crosses = "max"
    g1 += g2
    # Los datos viven en columnas ocultas: sin esto Excel no dibujaría nada.
    g1.visible_cells_only = False
    g2.visible_cells_only = False
    estilos = [(PAL["exito"], 28575, "solid"),
               ("9E9E9E", 12700, "dash"),
               ("1C4587", 19050, "solid")]
    for serie, (color, grosor, guion) in zip(g1.series, estilos):
        lp = LineProperties(solidFill=color, w=grosor)
        if guion == "dash":
            lp.prstDash = "dash"
        serie.graphicalProperties = GraphicalProperties(ln=lp)
        serie.smooth = True
        serie.marker = Marker(symbol="none")
    ws.add_chart(g1, "A46")

    for i in range(15, COL_NUM + N_SESIONES):
        ws.column_dimensions[get_column_letter(i)].hidden = True

    return ws


# ---------------------------------------------------------------------------
# PESTAÑA 3 — Calendario  (sin captura: construcción derivada, §8.3)
# ---------------------------------------------------------------------------

FILA_AGENDA = 12                      # fila de cabecera de la tabla AGENDA
N_AGENDA = N_TEMAS * N_SESIONES       # una fila por sesión posible
COL_MOTOR = 16                        # P: tema | Q: sesión | R: fecha | S: orden
COL_POS = 22                          # V: posición dentro del motor
COL_REJILLA = 7                       # G: primera columna de la rejilla mensual
FILA_REJILLA = 5      # fila 4 queda en blanco: separa la cabecera de días


def construir_calendario(wb):
    ws = wb.create_sheet("Calendario")
    ws.sheet_format.defaultRowHeight = ALTO_FILA_BASE
    anchos(ws, {"A": 22, "B": 18, "C": 14, "D": 14, "E": 14, "F": 3})
    for i in range(COL_REJILLA, COL_REJILLA + 7):
        ws.column_dimensions[get_column_letter(i)].width = 11
    fcab = f(PAL["txt_cab"])

    rej_prog = "Progreso!$B$2:$%s$%d" % (get_column_letter(1 + N_SESIONES), N_TEMAS + 1)
    num_prog = "Progreso!" + _rango_num()
    esp_prog = "Progreso!" + _rango_esp()

    # --- Semáforo de hoy ---------------------------------------------------
    pinta(ws, "A1", "Semáforo de hoy", PAL["verde_cab"], fcab, ALI_IZQ)
    pinta(ws, "B1", None, PAL["verde_cab"], fcab, ALI_IZQ)
    semaforo = [
        ("Hoy", "=TODAY()", FMT_FECHA),
        ("Sesiones para hoy", "=COUNTIF(%s,$B$2)" % rej_prog, "0"),
        ("Vencidas sin hacer",
         '=SUMPRODUCT((%s>0)*(%s<$B$2)*(%s=""))' % (num_prog, num_prog, esp_prog), "0"),
        ("Próximos 7 días",
         "=SUMPRODUCT((%s>$B$2)*(%s<=$B$2+VENTANA_PROXIMA))" % (num_prog, num_prog), "0"),
    ]
    for i, (etiqueta, formula, fmt) in enumerate(semaforo):
        r = 2 + i
        pinta(ws, "A%d" % r, etiqueta, PAL["res_cuerpo"], f(), ALI_IZQ)
        pinta(ws, "B%d" % r, formula, PAL["res_cuerpo"], f(bold=True), ALI_DER, fmt)
    alto(ws, range(1, 6))
    ws.conditional_formatting.add("B4", CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(name=FUENTE, size=TAM, bold=True, color=PAL["fallo"])))
    ws.conditional_formatting.add("B3", CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(name=FUENTE, size=TAM, bold=True, color=PAL["exito"])))

    # --- Rejilla mensual ---------------------------------------------------
    gc = get_column_letter(COL_REJILLA)
    gz = get_column_letter(COL_REJILLA + 6)
    pinta(ws, "%s1" % gc, "Rejilla mensual", PAL["verde_cab"], fcab, ALI_IZQ)
    for i in range(1, 7):
        pinta(ws, "%s1" % get_column_letter(COL_REJILLA + i), None,
              PAL["verde_cab"], fcab, ALI_IZQ)
    pinta(ws, "%s2" % gc, "=DATE(YEAR(HOY),MONTH(HOY),1)", PAL["res_cab"],
          f(bold=True), ALI_IZQ, "mmmm yyyy")
    for i in range(1, 7):
        pinta(ws, "%s2" % get_column_letter(COL_REJILLA + i), None,
              PAL["res_cab"], f(), ALI_IZQ)
    for j in range(7):
        pinta(ws, "%s3" % get_column_letter(COL_REJILLA + j),
              "=INDEX(DIAS_SEMANA,%d)" % (j + 1), PAL["res_cuerpo"],
              f(bold=True), ALI_CEN)
    # El ancla de ROW()/COLUMN() es la fila de nombres de día ($G$3), fuera de
    # la propia rejilla: si se anclara en la primera celda de la rejilla, cada
    # fórmula se referenciaría a sí misma.
    primer = ("$%s$2-WEEKDAY($%s$2,2)+1+(ROW()-ROW($%s$3)-%d)*DIAS_POR_SEMANA"
              "+(COLUMN()-COLUMN($%s$3))" % (gc, gc, gc, FILA_REJILLA - 3, gc))
    for i in range(6):
        for j in range(7):
            pinta(ws, "%s%d" % (get_column_letter(COL_REJILLA + j), FILA_REJILLA + i),
                  '=IF(MONTH(%s)<>MONTH($%s$2),"",%s)' % (primer, gc, primer),
                  "FFFFFF", f(), ALI_CEN, "d")
    alto(ws, range(1, 4))
    ws.row_dimensions[4].height = 4
    alto(ws, range(FILA_REJILLA, FILA_REJILLA + 6))
    rango_rej = "%s%d:%s%d" % (gc, FILA_REJILLA, gz, FILA_REJILLA + 5)
    ancla = "%s%d" % (gc, FILA_REJILLA)
    ws.conditional_formatting.add(rango_rej, FormulaRule(
        formula=['AND(%s<>"",COUNTIF(%s,%s)>=3)' % (ancla, rej_prog, ancla)],
        fill=relleno(PAL["escala_max"]), stopIfTrue=True))
    ws.conditional_formatting.add(rango_rej, FormulaRule(
        formula=['AND(%s<>"",COUNTIF(%s,%s)>=1)' % (ancla, rej_prog, ancla)],
        fill=relleno(PAL["escala_min"]), stopIfTrue=True))

    # --- Motor de ordenación (columnas ocultas) ---------------------------
    cm = {n: get_column_letter(COL_MOTOR + i)
          for i, n in enumerate(["tema", "ses", "fecha", "estado", "orden"])}
    for n, texto in (("tema", "Tema"), ("ses", "Sesión"), ("fecha", "Fecha"),
                     ("estado", "Estado marcado"), ("orden", "Orden cronológico")):
        pinta(ws, "%s1" % cm[n], texto, None, f(bold=True), ALI_IZQ)
    fin = N_AGENDA + 1
    for t in range(1, N_TEMAS + 1):
        for sesion in range(1, N_SESIONES + 1):
            r = 1 + (t - 1) * N_SESIONES + sesion
            pinta(ws, "%s%d" % (cm["tema"], r), t, None, f(), ALI_DER)
            pinta(ws, "%s%d" % (cm["ses"], r), sesion, None, f(), ALI_DER)
            pinta(ws, "%s%d" % (cm["fecha"], r),
                  "=INDEX(%s,$%s%d,$%s%d)" % (num_prog, cm["tema"], r, cm["ses"], r),
                  None, f(), ALI_DER)
            # INDEX sobre una celda vacía devuelve 0, no cadena vacía: se
            # normaliza aquí para que la agenda no muestre ceros.
            pinta(ws, "%s%d" % (cm["estado"], r),
                  '=IF(INDEX(%s,$%s%d,$%s%d)=0,"",INDEX(%s,$%s%d,$%s%d))'
                  % (esp_prog, cm["tema"], r, cm["ses"], r,
                     esp_prog, cm["tema"], r, cm["ses"], r),
                  None, f(), ALI_IZQ)
            pinta(ws, "%s%d" % (cm["orden"], r),
                  '=IF($%s%d=0,"",COUNTIFS($%s$2:$%s$%d,">0",$%s$2:$%s$%d,"<"&$%s%d)'
                  '+COUNTIFS($%s$2:%s%d,"="&$%s%d))'
                  % (cm["fecha"], r,
                     cm["fecha"], cm["fecha"], fin, cm["fecha"], cm["fecha"], fin,
                     cm["fecha"], r, cm["fecha"], cm["fecha"], r, cm["fecha"], r),
                  None, f(), ALI_DER)

    # --- Agenda cronológica ------------------------------------------------
    cabeceras = ["Fecha", "Día de la semana", "Tema", "Nº de sesión", "Estado"]
    for i, texto in enumerate(cabeceras):
        pinta(ws, "%s%d" % (get_column_letter(1 + i), FILA_AGENDA), texto,
              PAL["verde_cab"], fcab, ALI_IZQ)
    cpos = get_column_letter(COL_POS)
    pinta(ws, "%s%d" % (cpos, FILA_AGENDA), "Posición en el motor", None,
          f(bold=True), ALI_IZQ)
    r0 = FILA_AGENDA + 1
    for k in range(N_AGENDA):
        r = r0 + k
        banda = PAL["banda"] if (r - r0) % 2 else "FFFFFF"
        pinta(ws, "%s%d" % (cpos, r),
              '=IFERROR(MATCH(ROW()-ROW($%s$%d),$%s$2:$%s$%d,0),"")'
              % (cpos, FILA_AGENDA, cm["orden"], cm["orden"], fin), None, f(), ALI_DER)
        campos = [
            ("A", '=IF($%s%d="","",INDEX($%s$2:$%s$%d,$%s%d))'
             % (cpos, r, cm["fecha"], cm["fecha"], fin, cpos, r), FMT_FECHA, ALI_DER),
            ("B", '=IF($%s%d="","",INDEX(DIAS_SEMANA,WEEKDAY(INDEX($%s$2:$%s$%d,$%s%d),2)))'
             % (cpos, r, cm["fecha"], cm["fecha"], fin, cpos, r), None, ALI_IZQ),
            ("C", '=IF($%s%d="","","Tema "&INDEX($%s$2:$%s$%d,$%s%d))'
             % (cpos, r, cm["tema"], cm["tema"], fin, cpos, r), None, ALI_IZQ),
            ("D", '=IF($%s%d="","",INDEX($%s$2:$%s$%d,$%s%d))'
             % (cpos, r, cm["ses"], cm["ses"], fin, cpos, r), "0", ALI_DER),
            ("E", '=IF($%s%d="","",IF(INDEX($%s$2:$%s$%d,$%s%d)<>"",'
                  'INDEX($%s$2:$%s$%d,$%s%d),'
                  'IF(INDEX($%s$2:$%s$%d,$%s%d)<$B$2,"Vencida",'
                  'IF(INDEX($%s$2:$%s$%d,$%s%d)=$B$2,"Hoy","Programada"))))'
             % (cpos, r, cm["estado"], cm["estado"], fin, cpos, r,
                cm["estado"], cm["estado"], fin, cpos, r,
                cm["fecha"], cm["fecha"], fin, cpos, r,
                cm["fecha"], cm["fecha"], fin, cpos, r), None, ALI_IZQ),
        ]
        for col, formula, fmt, ali in campos:
            pinta(ws, "%s%d" % (col, r), formula, banda, f(), ali, fmt)
    alto(ws, range(FILA_AGENDA, r0 + N_AGENDA))
    tabla(ws, "AGENDA", "A%d:E%d" % (FILA_AGENDA, r0 + N_AGENDA - 1))
    ws.conditional_formatting.add("E%d:E%d" % (r0, r0 + N_AGENDA - 1), FormulaRule(
        formula=['$E%d="Vencida"' % r0], font=Font(name=FUENTE, size=TAM,
                                                   bold=True, color=PAL["fallo"])))
    ws.conditional_formatting.add("E%d:E%d" % (r0, r0 + N_AGENDA - 1), FormulaRule(
        formula=['$E%d="Hoy"' % r0], font=Font(name=FUENTE, size=TAM,
                                               bold=True, color=PAL["exito"])))

    for i in range(14, COL_POS + 1):
        ws.column_dimensions[get_column_letter(i)].hidden = True
    ws.freeze_panes = "A%d" % r0   # cabecera de la agenda siempre a la vista
    return ws


# ---------------------------------------------------------------------------
# PESTAÑA 4 — Preguntas
# ---------------------------------------------------------------------------

TIPOS = ["Referencia", "Analogía", "Evidencia"]
CHIP = {"Referencia": ("chip_ref_bg", "chip_ref_tx"),
        "Analogía":   ("chip_ana_bg", "chip_ana_tx"),
        "Evidencia":  ("chip_evi_bg", "chip_evi_tx")}

C4_MULTILINEA = (
    '1619: Alemania, contacto con nuevo sistema científico y matemático;\n'
    '1628: "Reglas para la dirección de la mente"\n'
    '1637: "Discurso del método"\n'
    '1641: "Meditaciones de la primera filosofía" + "Objeciones y respuestas"')

FILAS_PREGUNTAS = 21          # cabecera + 20 filas de trabajo


def construir_preguntas(wb):
    ws = wb.create_sheet("Preguntas")
    ws.sheet_format.defaultRowHeight = ALTO_FILA_BASE
    anchos(ws, {"A": 38.3, "B": 26.7, "C": 93.6})
    fcab = f(PAL["txt_cab"])

    pinta(ws, "A1", "Preguntas", PAL["verde_cab"], fcab, ALI_IZQ)
    pinta(ws, "B1", "Tipo de información", PAL["verde_cab"], fcab, ALI_IZQ)
    pinta(ws, "C1", "Respuesta", PAL["verde_cab"], fcab, ALI_IZQ)

    contenido = {
        2: ("Tema 1 - Descartes y 1ª Meditación", None, None),
        3: ("Nacimiento?", "Referencia", "1596 La Haye"),
        4: ("Pregunta 2", "Analogía", C4_MULTILINEA),
        5: ("Pregunta 3", "Evidencia", None),
        6: (None, "Referencia", None),
    }
    for r in range(2, FILAS_PREGUNTAS + 1):
        banda = PAL["banda"] if r % 2 else "FFFFFF"
        a, b, c = contenido.get(r, (None, None, None))
        pinta(ws, "A%d" % r, a, banda, f(bold=(r == 2)), ALI_IZQ)
        # el relleno del chip lo pone el formato condicional; aquí sólo la banda
        pinta(ws, "B%d" % r, b, banda, f(), ALI_IZQ)
        pinta(ws, "C%d" % r, c, banda, f(), ALI_ARR if r == 4 else ALI_IZQ)
    alto(ws, range(1, FILAS_PREGUNTAS + 1))
    ws.row_dimensions[4].height = 51.75      # cuatro líneas dentro de una celda

    tabla(ws, "PREGUNTAS", "A1:C%d" % FILAS_PREGUNTAS)

    rango_b = "B2:B%d" % FILAS_PREGUNTAS
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(TIPOS),
                        allow_blank=True, showErrorMessage=True,
                        errorTitle="Valor no permitido",
                        error="Elige una de las tres opciones: Referencia, "
                              "Analogía o Evidencia.",
                        promptTitle="Tipo de información",
                        prompt="Referencia, Analogía o Evidencia.")
    ws.add_data_validation(dv)
    dv.add(rango_b)

    # Equivalente en Excel de los chips de Google Sheets: formato condicional
    # sobre el mismo rango de la validación (ver SUPUESTOS.md).
    for tipo in TIPOS:
        bg, tx = CHIP[tipo]
        ws.conditional_formatting.add(rango_b, FormulaRule(
            formula=['$B2="%s"' % tipo], fill=relleno(PAL[bg]),
            font=Font(name=FUENTE, size=TAM, color=PAL[tx]), stopIfTrue=True))
    # píldora gris del desplegable vacío (ref/01)
    ws.conditional_formatting.add(rango_b, FormulaRule(
        formula=['$B2=""'], fill=relleno(PAL["pildora"]), stopIfTrue=True))
    return ws


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def construir(destino):
    wb = Workbook()
    wb.remove(wb.active)
    # Fuente por defecto de todo el libro: sustituir la fuente 0 de la tabla de
    # estilos es lo único que Excel lee como «Normal» (tocar el NamedStyle no
    # llega a styles.xml).
    wb._fonts[0] = Font(name=FUENTE, size=TAM)
    wb._named_styles["Normal"].font = Font(name=FUENTE, size=TAM)
    # Metadatos fijos: sin esto, la marca de tiempo cambiaría en cada ejecución
    # y el libro no sería idempotente byte a byte.
    wb.properties.creator = "Proyecto ATENEO — build/construir.py"
    wb.properties.lastModifiedBy = wb.properties.creator
    wb.properties.title = "Ejemplo Planificador Oposición"
    wb.properties.created = dt.datetime(2026, 1, 1, 0, 0, 0)
    wb.properties.modified = dt.datetime(2026, 1, 1, 0, 0, 0)

    construir_estrategia(wb)
    construir_progreso(wb)
    construir_calendario(wb)
    construir_preguntas(wb)

    for i, (_etiqueta, _horas, nombre) in enumerate(TARIFAS):
        nombre_rango(wb, nombre, "Estrategia!$C$%d" % (3 + i))
    nombre_rango(wb, "DIAS_ESTUDIO", "Progreso!$M$2:$M$5")
    nombre_rango(wb, "INTERVALOS", "Progreso!$M$9:$M$17")
    nombre_rango(wb, "DIAS_SEMANA", "Progreso!$O$2:$O$8")
    nombre_rango(wb, "DESPLAZAMIENTO", "Progreso!$P$2:$P$8")
    nombre_rango(wb, "ESTADO_EXITO", "Progreso!$AN$1")
    nombre_rango(wb, "ESTADO_FALLO", "Progreso!$AN$2")
    nombre_rango(wb, "DIAS_POR_SEMANA", "Progreso!$AN$4")
    nombre_rango(wb, "VENTANA_PROXIMA", "Progreso!$AN$5")
    nombre_rango(wb, "HOY", "Calendario!$B$2")

    wb.save(destino)
    _normalizar_zip(destino)
    return destino


SELLO = (2026, 1, 1, 0, 0, 0)


def _normalizar_zip(ruta):
    """Reescribe el .xlsx con marcas de tiempo fijas.

    openpyxl sella cada entrada del ZIP y la propiedad `modified` con la hora
    de guardado; sin normalizarlas, dos ejecuciones del script darían archivos
    distintos byte a byte aunque el contenido fuese idéntico.
    """
    with zipfile.ZipFile(ruta) as z:
        partes = [(info, z.read(info.filename)) for info in z.infolist()]
    temporal = ruta + ".tmp"
    with zipfile.ZipFile(temporal, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as out:
        for info, datos in partes:
            if info.filename == "docProps/core.xml":
                datos = re.sub(
                    b"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                    b"2026-01-01T00:00:00Z</dcterms:modified>", datos)
            nuevo = zipfile.ZipInfo(info.filename, date_time=SELLO)
            nuevo.compress_type = info.compress_type
            nuevo.external_attr = info.external_attr
            out.writestr(nuevo, datos)
    os.replace(temporal, ruta)


if __name__ == "__main__":
    salida = (sys.argv[1] if len(sys.argv) > 1 else
              os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           "Ejemplo Planificador Oposición.xlsx"))
    construir(salida)
    print("Escrito: %s" % salida)
