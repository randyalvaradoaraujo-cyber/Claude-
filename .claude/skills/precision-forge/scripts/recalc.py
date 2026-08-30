#!/usr/bin/env python3
"""Precision Forge recalculator.

Computes every formula in a workbook with a pure-Python engine, so the numbers
can be verified without Excel and without trusting whatever values happen to be
cached in the file.

This matters more than it sounds. A workbook written by a library carries no
computed values at all -- every formula is an untested claim until something
evaluates it. And a workbook edited by hand can carry values that no longer
match its formulas, which is how a "reviewed" model ships with a stale number
in it. Both cases are invisible to the eye and obvious to this script.

Three things it answers:
  1. Does every formula evaluate without an error?
  2. Do the computed values match the values stored in the file? (drift)
  3. Do they match what you said they should be? (golden test, --expect)

Exit codes:  0 = clean, 1 = errors/mismatches found, 2 = could not evaluate.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import warnings

warnings.filterwarnings("ignore")

CELL_KEY = re.compile(r"^'?\[(?P<file>[^\]]+)\](?P<sheet>.+?)'?!(?P<cell>\$?[A-Z]{1,3}\$?\d{1,7})$")


def scalar(val):
    """Pull a single Python value out of whatever the engine returns.

    Engine cells come back as range objects wrapping numpy arrays; a 1x1 array
    is a plain cell. Anything larger is a spilled range with no single value to
    report, so it is skipped.
    """
    v = getattr(val, "value", val)
    if type(v).__name__ == "XlError":
        return v
    if hasattr(v, "ndim"):
        if getattr(v, "size", 0) != 1:
            return None
        if v.ndim:
            v = v.reshape(-1)[0]
        if type(v).__name__ == "XlError":
            return v
    if hasattr(v, "item"):
        try:
            v = v.item()
        except Exception:
            pass
    return v


def is_error(v):
    return v is not None and type(v).__name__ == "XlError"


def compute(path):
    try:
        import formulas
    except ImportError:
        sys.stderr.write(
            "precision-forge: the recalculation engine is missing (pip install formulas).\n"
            "Without it, no value-level verification runs -- report that rather than\n"
            "assuming the numbers are right.\n")
        raise SystemExit(2)
    model = formulas.ExcelModel().loads(path).finish()
    solution = model.calculate()
    out = {}
    for key, val in solution.items():
        m = CELL_KEY.match(str(key))
        if not m:
            continue
        sheet = m.group("sheet").strip("'").upper()
        cell = m.group("cell").replace("$", "").upper()
        v = scalar(val)
        if v is None and not is_error(val):
            continue
        out["%s!%s" % (sheet, cell)] = v
    return out


def close_enough(a, b, tol):
    if isinstance(a, bool) or isinstance(b, bool):
        return a == b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if math.isnan(a) and math.isnan(b):
            return True
        return abs(a - b) <= max(tol, tol * max(abs(a), abs(b)))
    if isinstance(a, str) and isinstance(b, str):
        return a.strip() == b.strip()
    return a == b


def main():
    ap = argparse.ArgumentParser(description="Precision Forge recalculator")
    ap.add_argument("workbook")
    ap.add_argument("--json", dest="json_out", help="write computed values + findings here")
    ap.add_argument("--expect", help="JSON of {'Sheet!A1': expected_value} to assert against")
    ap.add_argument("--compare-cached", action="store_true",
                    help="compare computed values against the values stored in the file")
    ap.add_argument("--tolerance", type=float, default=1e-9)
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        sys.stderr.write("precision-forge: no such file: %s\n" % args.workbook)
        return 2
    try:
        values = compute(args.workbook)
    except SystemExit:
        raise
    except Exception as exc:
        sys.stderr.write("precision-forge: recalculation failed: %s\n" % exc)
        return 2

    findings = []
    for ref, v in sorted(values.items()):
        if is_error(v):
            sheet, cell = ref.split("!", 1)
            findings.append({
                "check": "CALC_ERROR", "severity": "critical", "sheet": sheet, "cell": cell,
                "signature": "CALC_ERROR::%s" % v,
                "message": "Formula evaluates to %s." % v,
                "detail": "", "fix": "Repair the precedent that produces the error.",
            })

    if args.compare_cached:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(args.workbook, data_only=True)
            wbf = openpyxl.load_workbook(args.workbook, data_only=False)
            for ws in wbf.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if not (isinstance(c.value, str) and c.value.startswith("=")):
                            continue
                        cached = wb[ws.title][c.coordinate].value
                        if cached is None:
                            continue
                        got = values.get("%s!%s" % (ws.title.upper(), c.coordinate))
                        if got is None or is_error(got):
                            continue
                        if not close_enough(got, cached, args.tolerance):
                            findings.append({
                                "check": "VALUE_DRIFT", "severity": "critical",
                                "sheet": ws.title, "cell": c.coordinate,
                                "signature": "VALUE_DRIFT",
                                "message": "Stored value %r does not match the recomputed %r."
                                           % (cached, got),
                                "detail": c.value,
                                "fix": "The saved value is stale or was typed over the formula. "
                                       "Recalculate and save, or restore the formula.",
                            })
        except Exception as exc:
            sys.stderr.write("precision-forge: cached comparison skipped: %s\n" % exc)

    if args.expect:
        with open(args.expect, encoding="utf-8") as fh:
            expected = json.load(fh)
        for ref, want in expected.items():
            key = ref.upper().replace("$", "")
            got = values.get(key)
            if got is None:
                findings.append({
                    "check": "EXPECT_MISSING", "severity": "critical",
                    "sheet": key.split("!")[0], "cell": key.split("!")[-1],
                    "signature": "EXPECT_MISSING",
                    "message": "Expected %r at %s but the cell produced nothing." % (want, ref),
                    "detail": "", "fix": "Check the reference in your expectations file.",
                })
            elif not close_enough(got, want, args.tolerance):
                findings.append({
                    "check": "EXPECT_MISMATCH", "severity": "critical",
                    "sheet": key.split("!")[0], "cell": key.split("!")[-1],
                    "signature": "EXPECT_MISMATCH",
                    "message": "%s = %r, expected %r." % (ref, got, want),
                    "detail": "", "fix": "Either the model is wrong or the expectation is. "
                                         "Resolve which before moving on -- do not adjust the "
                                         "expectation just to make it pass.",
                })

    report = {
        "file": os.path.abspath(args.workbook),
        "cells_computed": len(values),
        "findings": findings,
        "values": {k: (str(v) if is_error(v) else v) for k, v in values.items()},
    }
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as fh:
            json.dump(report, fh, indent=2, ensure_ascii=False, default=str)

    if not args.quiet:
        print("PRECISION FORGE recalc  %s" % report["file"])
        print("  %d cells computed, %d finding(s)" % (len(values), len(findings)))
        for f in findings:
            print("\n  %-8s %-15s %s!%s" % (f["severity"].upper(), f["check"], f["sheet"], f["cell"]))
            print("      %s" % f["message"])
            if f["fix"]:
                print("      fix: %s" % f["fix"])
        if not findings:
            print("  Every formula evaluates cleanly.")

    return 1 if findings else 0


if __name__ == "__main__":
    sys.exit(main())
