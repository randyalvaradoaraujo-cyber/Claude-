#!/usr/bin/env python3
"""Precision Forge auditor.

Reads a workbook twice (formulas and cached values), runs a catalog of checks,
and emits findings as JSON, text, or both.

The point of this script is to make spreadsheet defects *mechanically visible*.
A human reading a formula column sees what they expect to see; a diff of R1C1
patterns sees the one row where the pattern broke. Run it after every build
slice, not once at the end -- a defect caught while you still remember writing
the block costs a minute, the same defect found three blocks later costs an
hour of bisecting.

Exit codes:  0 = clean at the --fail-on threshold, 1 = findings at/above it,
             2 = the workbook could not be read.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.stderr.write("precision-forge: openpyxl is required (pip install openpyxl)\n")
    sys.exit(2)

SEVERITIES = ["critical", "high", "medium", "low", "info"]
SEV_RANK = {s: i for i, s in enumerate(SEVERITIES)}

ERROR_LITERALS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!")
VOLATILE_FUNCS = ("NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT", "CELL", "INFO")

FIXES = {
    "ERROR_LITERAL": "Trace the precedent that produces the error and repair it at the source. Wrapping it in IFERROR hides a real defect unless the error is a documented, expected state.",
    "BAD_SHEET_REF": "The referenced sheet does not exist in this workbook. Fix the sheet name (watch for renames and trailing spaces) or create the sheet.",
    "BROKEN_NAME": "Redefine the named range. Names that survive a deleted sheet keep pointing at #REF! and poison every formula that uses them.",
    "SELF_REF": "A cell referencing itself is a circular reference. Restructure so the value comes from a precedent cell.",
    "VLOOKUP_COL_OUT_OF_RANGE": "col_index_num is larger than the width of table_array, so this returns #REF! for every row. Widen the range or lower the index.",
    "VLOOKUP_APPROX": "Add FALSE (or 0) as the fourth argument. Without it VLOOKUP does an approximate match and silently returns the wrong row when the key is not found in a sorted-ascending column.",
    "VLOOKUP_UNANCHORED": "Anchor table_array with $ (e.g. $B$2:$D$500). Unanchored, the lookup range slides as the formula is filled down and rows drop out of the search.",
    "PATTERN_BREAK": "This formula differs from the pattern its neighbours follow. Either it is the bug, or it is a deliberate exception that deserves a comment and a note in the handoff.",
    "CONSTANT_IN_FORMULA_RUN": "A pasted value sits inside a block of formulas. It will not update when its inputs change. Restore the formula, or move the input out to a labelled input cell.",
    "NUMBER_AS_TEXT": "Numbers stored as text are skipped by SUM and never match numeric lookups, so totals are silently low. Convert to real numbers.",
    "LOOKUP_KEY_DUPLICATE": "The first column of this lookup range has duplicate keys. VLOOKUP returns the first hit only, so some rows silently take the wrong record. De-duplicate or use a compound key.",
    "TRAILING_SPACE": "Leading/trailing whitespace makes exact-match lookups fail on values that look identical on screen. TRIM the source.",
    "EXTERNAL_LINK": "This formula points at another workbook. It breaks the moment the file moves and it cannot be recalculated in isolation. Import the data or document the dependency.",
    "VOLATILE": "Volatile functions recalculate on every change and make the workbook non-deterministic, so the same inputs can produce different saved values. Freeze the value or isolate it in one documented cell.",
    "WHOLE_COLUMN_REF": "Whole-column references scan every row of the sheet. Bound the range to the data, or use a table reference.",
    "MERGED_IN_DATA": "Merged cells inside a data region break sorting, filtering and structured references. Use 'Center Across Selection' for the visual effect instead.",
    "IFERROR_BLANK_MASK": "IFERROR returning blank hides every failure mode, including the ones you did not anticipate. Return a sentinel you can filter on, or handle the specific error.",
    "NARROW_COLUMN": "The column is likely too narrow for its widest value, which renders as ###. Widen it so printed and shared copies stay readable.",
    "TOTAL_RANGE_MISMATCH": "The aggregate at the edge of this block does not cover the whole block. This is what happens when rows are added after the total was written: the total keeps working, it is just quietly wrong. Extend the range to the full block.",
    "NO_CACHED_VALUES": "This workbook carries no computed values (normal for a file just written by a library). Run recalc.py to compute them before trusting any value-level check.",
}

# --------------------------------------------------------------------------
# formula parsing helpers
# --------------------------------------------------------------------------

_STR_RE = re.compile(r'"(?:[^"]|"")*"')
_REF_RE = re.compile(
    r"""(?<![A-Za-z0-9_.\x01])
        (?:(?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?
        (?P<ac>\$?)(?P<col>[A-Za-z]{1,3})(?P<ar>\$?)(?P<row>\d{1,7})
        (?![A-Za-z0-9_(\x01])""",
    re.X,
)
_RANGE_RE = re.compile(r"\$?[A-Za-z]{1,3}\$?\d{1,7}:\$?[A-Za-z]{1,3}\$?\d{1,7}")
_AGG_RE = re.compile(
    r"^=\s*(SUM|AVERAGE|COUNT|COUNTA|MIN|MAX|MEDIAN|PRODUCT|STDEV\.?[SP]?|VAR\.?[SP]?|SUBTOTAL|AGGREGATE)\s*\(",
    re.I)
_FULLRANGE_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?(?P<c1>[A-Za-z]{1,3})\$?(?P<r1>\d{1,7}):\$?(?P<c2>[A-Za-z]{1,3})\$?(?P<r2>\d{1,7})")
_WHOLE_COL_RE = re.compile(r"(?<![A-Za-z0-9_$])\$?[A-Za-z]{1,3}:\$?[A-Za-z]{1,3}(?![A-Za-z0-9_])")


def mask_strings(formula: str):
    """Replace string literals with markers so they never confuse ref parsing."""
    lits = []

    def _sub(m):
        lits.append(m.group(0))
        return "\x01%d\x01" % (len(lits) - 1)

    return _STR_RE.sub(_sub, formula), lits


def normalize(formula: str, row: int, col: int) -> str:
    """Return an R1C1-style normal form.

    Two cells that were filled from the same source formula normalize to the
    same string, which is what lets us spot the one row that was edited by hand.
    String literals collapse to a placeholder (per-row text is usually
    legitimate); numeric literals are kept, because a changed constant in one
    row of a filled column is exactly the defect we are hunting.
    """
    masked, _ = mask_strings(formula)

    def _sub(m):
        sheet = ("%s!" % m.group("sheet")) if m.group("sheet") else ""
        c = column_index_from_string(m.group("col").upper())
        r = int(m.group("row"))
        cpart = "C%d" % c if m.group("ac") else "C[%d]" % (c - col)
        rpart = "R%d" % r if m.group("ar") else "R[%d]" % (r - row)
        return "%s%s%s" % (sheet, rpart, cpart)

    out = _REF_RE.sub(_sub, masked)
    return re.sub(r"\x01\d+\x01", '"@"', out).upper().replace(" ", "")


def split_args(argstr: str):
    """Split a function argument list on top-level commas."""
    args, depth, cur, in_str = [], 0, [], False
    i = 0
    while i < len(argstr):
        ch = argstr[i]
        if in_str:
            cur.append(ch)
            if ch == '"':
                if i + 1 < len(argstr) and argstr[i + 1] == '"':
                    cur.append('"')
                    i += 1
                else:
                    in_str = False
        elif ch == '"':
            in_str = True
            cur.append(ch)
        elif ch in "([":
            depth += 1
            cur.append(ch)
        elif ch in ")]":
            depth -= 1
            cur.append(ch)
        elif ch == "," and depth == 0:
            args.append("".join(cur).strip())
            cur = []
        else:
            cur.append(ch)
        i += 1
    args.append("".join(cur).strip())
    return args


def find_calls(formula: str, name: str):
    """Yield the argument list of every call to `name` in the formula."""
    pat = re.compile(r"(?<![A-Za-z0-9_.])" + name + r"\s*\(", re.I)
    for m in pat.finditer(formula):
        depth, i = 1, m.end()
        while i < len(formula) and depth:
            if formula[i] == "(":
                depth += 1
            elif formula[i] == ")":
                depth -= 1
            i += 1
        yield split_args(formula[m.end(): i - 1])


def range_width(rng: str):
    m = _RANGE_RE.search(rng)
    if not m:
        return None
    a, b = m.group(0).split(":")
    ca = column_index_from_string(re.sub(r"[^A-Za-z]", "", a).upper())
    cb = column_index_from_string(re.sub(r"[^A-Za-z]", "", b).upper())
    return abs(cb - ca) + 1


# --------------------------------------------------------------------------
# audit
# --------------------------------------------------------------------------

class Audit:
    def __init__(self, path, ledger=None, max_per_check=40):
        self.path = path
        self.ledger = ledger or {}
        self.max_per_check = max_per_check
        self.findings = []
        self._counts = Counter()
        self.muted_hits = 0
        self.muted = set(self.ledger.get("muted", []))

    def add(self, check, severity, sheet, cell, message, detail="", signature=None):
        sig = signature or check
        full_sig = "%s::%s" % (check, sig) if not sig.startswith(check) else sig
        if full_sig in self.muted:
            self.muted_hits += 1
            return
        self._counts[check] += 1
        if self._counts[check] > self.max_per_check:
            if self._counts[check] == self.max_per_check + 1:
                self.findings.append({
                    "check": check, "severity": "info", "sheet": sheet, "cell": "",
                    "signature": "%s::truncated" % check,
                    "message": "More than %d findings for %s; showing the first %d. Fix these and re-run."
                               % (self.max_per_check, check, self.max_per_check),
                    "detail": "", "fix": "",
                })
            return
        self.findings.append({
            "check": check, "severity": severity, "sheet": sheet, "cell": cell,
            "signature": full_sig, "message": message, "detail": detail,
            "fix": FIXES.get(check, ""),
        })

    # -- per-cell formula checks ------------------------------------------
    def check_formula(self, ws, cell, sheetnames):
        f = cell.value
        if not isinstance(f, str) or not f.startswith("="):
            return
        coord, sheet = cell.coordinate, ws.title
        up = f.upper()

        for lit in ERROR_LITERALS:
            if lit in up:
                self.add("ERROR_LITERAL", "critical", sheet, coord,
                         "Formula contains the literal error %s." % lit, f, "ERROR_LITERAL::%s" % lit)

        masked, _ = mask_strings(f)
        for m in _REF_RE.finditer(masked):
            sh = m.group("sheet")
            if sh:
                clean = sh.strip("'")
                if clean not in sheetnames:
                    self.add("BAD_SHEET_REF", "critical", sheet, coord,
                             "References sheet '%s', which does not exist." % clean, f,
                             "BAD_SHEET_REF::%s" % clean)
            elif m.group("col").upper() + m.group("row") == coord:
                self.add("SELF_REF", "critical", sheet, coord,
                         "Formula references its own cell (circular).", f)

        if re.search(r"\[\d+\]|\[[^\]]*\.xls", f, re.I):
            self.add("EXTERNAL_LINK", "medium", sheet, coord,
                     "Formula links to an external workbook.", f)

        for fn in VOLATILE_FUNCS:
            if re.search(r"(?<![A-Za-z0-9_.])" + fn + r"\s*\(", up):
                self.add("VOLATILE", "medium", sheet, coord,
                         "Uses the volatile function %s()." % fn, f, "VOLATILE::%s" % fn)

        if _WHOLE_COL_RE.search(masked):
            self.add("WHOLE_COLUMN_REF", "medium", sheet, coord,
                     "Uses a whole-column reference.", f)

        if re.search(r"IFERROR\s*\([^,]*,\s*\"\"\s*\)", up):
            self.add("IFERROR_BLANK_MASK", "low", sheet, coord,
                     "IFERROR silently returns blank on any failure.", f)

        for fname in ("VLOOKUP", "HLOOKUP"):
            for args in find_calls(f, fname):
                if len(args) < 4:
                    self.add("VLOOKUP_APPROX", "critical", sheet, coord,
                             "%s has no range_lookup argument, so it does an approximate match." % fname,
                             f, "VLOOKUP_APPROX::missing")
                elif args[3].strip().upper() in ("TRUE", "1"):
                    self.add("VLOOKUP_APPROX", "high", sheet, coord,
                             "%s uses approximate match (TRUE). Correct only on a sorted-ascending key." % fname,
                             f, "VLOOKUP_APPROX::true")
                if len(args) >= 2:
                    ta = args[1]
                    if _RANGE_RE.search(ta) and "$" not in ta:
                        self.add("VLOOKUP_UNANCHORED", "high", sheet, coord,
                                 "%s table_array '%s' is not anchored with $." % (fname, ta), f)
                    w = range_width(ta)
                    if w and len(args) >= 3 and re.fullmatch(r"\d+", args[2].strip()):
                        idx = int(args[2].strip())
                        if idx > w:
                            self.add("VLOOKUP_COL_OUT_OF_RANGE", "critical", sheet, coord,
                                     "col_index_num %d exceeds the %d-column table_array '%s'." % (idx, w, ta), f)

        for rid, rule in (self.ledger.get("rules") or {}).items():
            try:
                if re.search(rule["pattern"], f, re.I):
                    self.add(rid, rule.get("severity", "medium"), sheet, coord,
                             rule.get("message", "Matches learned rule %s." % rid), f)
            except re.error:
                pass

    # -- block-level consistency ------------------------------------------
    def check_blocks(self, ws, axis):
        """Find runs of adjacent cells that behave as one calculated block and
        flag the members that break the block's pattern."""
        groups = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or (isinstance(c.value, str) and not c.value.strip()):
                    continue
                key, pos = (c.column, c.row) if axis == "col" else (c.row, c.column)
                groups[key].append((pos, c))

        for key, cells in groups.items():
            cells.sort(key=lambda t: t[0])
            run = []
            prev = None
            for pos, c in cells + [(None, None)]:
                if prev is not None and pos is not None and pos == prev + 1:
                    run.append((pos, c))
                else:
                    self._eval_run(ws, run, axis)
                    run = [(pos, c)] if c is not None else []
                prev = pos

    def _eval_run(self, ws, run, axis):
        if len(run) < 4:
            return
        # A leading text cell is a header, not a member of the block.
        if isinstance(run[0][1].value, str) and not str(run[0][1].value).startswith("="):
            run = run[1:]
        if len(run) < 4:
            return
        norms, formulas_n = {}, 0
        for pos, c in run:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                formulas_n += 1
                norms[c.coordinate] = normalize(v, c.row, c.column)
        if formulas_n < 3 or formulas_n / len(run) < 0.7:
            return
        mode, mode_n = Counter(norms.values()).most_common(1)[0]
        if mode_n / formulas_n < 0.7:
            return  # no single dominant pattern; not a filled block
        span = "%s:%s" % (run[0][1].coordinate, run[-1][1].coordinate)
        edges = {run[0][1].coordinate, run[-1][1].coordinate}
        for pos, c in run:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                if norms[c.coordinate] != mode:
                    # A total sitting at the edge of a filled block is normal.
                    # What is worth checking there is whether its range still
                    # covers the block -- that is the defect rows-added-later
                    # introduces, and it never announces itself.
                    if c.coordinate in edges and _AGG_RE.match(v):
                        self._check_total(ws, run, c, axis, span)
                        continue
                    self.add("PATTERN_BREAK", "high", ws.title, c.coordinate,
                             "Breaks the pattern shared by %d of %d formulas in %s (%s)."
                             % (mode_n, formulas_n, span, axis),
                             "%s   |   block pattern: %s" % (v, mode),
                             "PATTERN_BREAK::%s" % mode)
            else:
                self.add("CONSTANT_IN_FORMULA_RUN", "high", ws.title, c.coordinate,
                         "Hard value inside the calculated block %s." % span,
                         "value=%r   |   block pattern: %s" % (v, mode),
                         "CONSTANT_IN_FORMULA_RUN::%s" % mode)

    def _check_total(self, ws, run, cell, axis, span):
        """Verify that an edge aggregate spans the block it summarises."""
        members = [c for _, c in run if c.coordinate != cell.coordinate]
        if not members:
            return
        m = _FULLRANGE_RE.search(cell.value)
        if not m:
            return
        if axis == "col":
            lo, hi = int(m.group("r1")), int(m.group("r2"))
            want_lo = min(c.row for c in members)
            want_hi = max(c.row for c in members)
            unit = "rows"
        else:
            lo = column_index_from_string(m.group("c1").upper())
            hi = column_index_from_string(m.group("c2").upper())
            want_lo = min(c.column for c in members)
            want_hi = max(c.column for c in members)
            unit = "columns"
        lo, hi = min(lo, hi), max(lo, hi)
        if lo > want_lo or hi < want_hi:
            missing = (want_lo - lo if lo > want_lo else 0) + (want_hi - hi if hi < want_hi else 0)
            self.add("TOTAL_RANGE_MISMATCH", "high", ws.title, cell.coordinate,
                     "Aggregate covers %s %d-%d but the block %s spans %s %d-%d, leaving %d out."
                     % (unit, lo, hi, span, unit, want_lo, want_hi, missing),
                     cell.value)

    # -- data hygiene ------------------------------------------------------
    def check_data(self, ws):
        numlike = re.compile(r"^\s*-?[\d]{1,3}(?:[.,\s]?\d{3})*(?:[.,]\d+)?\s*%?\s*$")
        cols = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v != v.strip() and v.strip():
                    self.add("TRAILING_SPACE", "medium", ws.title, c.coordinate,
                             "Text has leading/trailing whitespace.", repr(v))
                cols[c.column].append(c)

        for col, cells in cols.items():
            nums = [c for c in cells if isinstance(c.value, (int, float)) and not isinstance(c.value, bool)]
            texts = [c for c in cells if isinstance(c.value, str) and not c.value.startswith("=")
                     and numlike.match(c.value)]
            if len(nums) >= 3 and texts:
                for c in texts:
                    self.add("NUMBER_AS_TEXT", "high", ws.title, c.coordinate,
                             "Numeric-looking text in column %s, which is otherwise numeric."
                             % get_column_letter(col), repr(c.value))

        for rng in list(ws.merged_cells.ranges):
            if rng.min_row > 1 and (rng.max_row - rng.min_row) >= 1:
                self.add("MERGED_IN_DATA", "medium", ws.title, str(rng),
                         "Merged range spans data rows.", str(rng))

        for letter, dim in (ws.column_dimensions or {}).items():
            if not dim.width:
                continue
            try:
                idx = column_index_from_string(letter)
            except ValueError:
                continue
            widest = 0
            for c in cols.get(idx, []):
                if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    widest = max(widest, len(("%.2f" % c.value).rstrip("0").rstrip(".")))
            if widest and dim.width < widest + 1:
                self.add("NARROW_COLUMN", "low", ws.title, "%s1" % letter,
                         "Column %s is %.1f wide but holds values up to %d characters."
                         % (letter, dim.width, widest), "")

    def check_lookup_keys(self, wb, sheetnames):
        """Duplicate keys are only a defect if something actually looks them up,
        so derive the key columns from the VLOOKUP ranges themselves."""
        keyranges = set()
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if not (isinstance(c.value, str) and c.value.startswith("=")):
                        continue
                    for args in find_calls(c.value, "VLOOKUP"):
                        if len(args) < 2:
                            continue
                        ta = args[1]
                        m = _RANGE_RE.search(ta)
                        if not m:
                            continue
                        sheet = ws.title
                        sm = re.search(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!", ta)
                        if sm:
                            sheet = sm.group(1) or sm.group(2)
                        if sheet in sheetnames:
                            # Normalise anchors so $A$2:$C$4 and A2:C4 are one range.
                            keyranges.add((sheet, m.group(0).replace("$", "").upper()))
        for sheet, rng in sorted(keyranges):
            ws = wb[sheet]
            a, b = rng.split(":")
            col = column_index_from_string(re.sub(r"[^A-Za-z]", "", a).upper())
            r1 = int(re.sub(r"[^\d]", "", a))
            r2 = int(re.sub(r"[^\d]", "", b))
            seen, dupes = {}, {}
            for r in range(min(r1, r2), min(max(r1, r2), ws.max_row) + 1):
                v = ws.cell(r, col).value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                k = v.strip().lower() if isinstance(v, str) else v
                if k in seen:
                    dupes.setdefault(k, [seen[k]]).append(r)
                else:
                    seen[k] = r
            for k, rows in list(dupes.items())[:10]:
                self.add("LOOKUP_KEY_DUPLICATE", "high", sheet,
                         "%s%d" % (get_column_letter(col), rows[1]),
                         "Lookup key %r appears in rows %s of %s." % (k, rows, rng), "")

    def check_names(self, wb):
        try:
            items = list(wb.defined_names.items())
        except Exception:
            return
        for name, dn in items:
            dest = str(getattr(dn, "attr_text", "") or "")
            if "#REF" in dest:
                self.add("BROKEN_NAME", "critical", "", "",
                         "Defined name '%s' points at #REF!." % name, dest, "BROKEN_NAME::%s" % name)
            else:
                m = re.match(r"^(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!", dest)
                if m:
                    sh = m.group(1) or m.group(2)
                    if sh not in wb.sheetnames:
                        self.add("BROKEN_NAME", "critical", "", "",
                                 "Defined name '%s' points at missing sheet '%s'." % (name, sh), dest,
                                 "BROKEN_NAME::%s" % name)

    def run(self):
        wbf = openpyxl.load_workbook(self.path, data_only=False)
        try:
            wbv = openpyxl.load_workbook(self.path, data_only=True)
        except Exception:
            wbv = None
        sheetnames = set(wbf.sheetnames)

        has_formula = has_cached = False
        for ws in wbf.worksheets:
            wsv = wbv[ws.title] if wbv else None
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        has_formula = True
                        if wsv is not None and wsv[c.coordinate].value is not None:
                            has_cached = True
                    self.check_formula(ws, c, sheetnames)
            for axis in ("col", "row"):
                self.check_blocks(ws, axis)
            self.check_data(ws)

        if wbv is not None:
            for ws in wbv.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if isinstance(c.value, str) and c.value.strip() in ERROR_LITERALS:
                            self.add("ERROR_LITERAL", "critical", ws.title, c.coordinate,
                                     "Cached value is the error %s." % c.value.strip(), c.value,
                                     "ERROR_LITERAL::%s" % c.value.strip())

        self.check_lookup_keys(wbf, sheetnames)
        self.check_names(wbf)

        if has_formula and not has_cached:
            self.add("NO_CACHED_VALUES", "info", "", "",
                     "Workbook has formulas but no computed values; value-level checks did not run.", "")
        return self.findings


def main():
    ap = argparse.ArgumentParser(description="Precision Forge auditor")
    ap.add_argument("workbook")
    ap.add_argument("--ledger", help="path to ledger.json (mutes + learned rules)")
    ap.add_argument("--json", dest="json_out", help="write findings JSON here")
    ap.add_argument("--format", choices=["text", "json", "both"], default="text")
    ap.add_argument("--fail-on", choices=SEVERITIES + ["none"], default="high",
                    help="exit 1 if any finding is at or above this severity")
    ap.add_argument("--max-per-check", type=int, default=40)
    args = ap.parse_args()

    ledger = {}
    if args.ledger and os.path.exists(args.ledger):
        with open(args.ledger, encoding="utf-8") as fh:
            ledger = json.load(fh)

    try:
        audit = Audit(args.workbook, ledger, args.max_per_check)
        findings = audit.run()
    except Exception as exc:
        sys.stderr.write("precision-forge: cannot read %s: %s\n" % (args.workbook, exc))
        return 2

    findings.sort(key=lambda f: (SEV_RANK[f["severity"]], f["check"], f["sheet"], f["cell"]))
    for i, f in enumerate(findings, 1):
        f["id"] = "F%03d" % i
    summary = Counter(f["severity"] for f in findings)
    report = {
        "file": os.path.abspath(args.workbook),
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "summary": {s: summary.get(s, 0) for s in SEVERITIES},
        "muted_suppressed": audit.muted_hits,
        "findings": findings,
    }

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as fh:
            json.dump(report, fh, indent=2, ensure_ascii=False)

    if args.format in ("json", "both"):
        print(json.dumps(report, indent=2, ensure_ascii=False))
    if args.format in ("text", "both"):
        print("PRECISION FORGE  %s" % report["file"])
        print("  " + "  ".join("%s=%d" % (s, report["summary"][s]) for s in SEVERITIES))
        if audit.muted_hits:
            print("  %d finding(s) suppressed by the ledger" % audit.muted_hits)
        if not findings:
            print("\n  No findings. Proceed to the next slice.")
        for f in findings:
            loc = "%s!%s" % (f["sheet"], f["cell"]) if f["sheet"] else (f["cell"] or "-")
            print("\n[%s] %-8s %-26s %s" % (f["id"], f["severity"].upper(), f["check"], loc))
            print("     %s" % f["message"])
            if f["detail"]:
                print("     %s" % f["detail"])
            if f["fix"]:
                print("     fix: %s" % f["fix"])

    if args.fail_on != "none":
        thr = SEV_RANK[args.fail_on]
        if any(SEV_RANK[f["severity"]] <= thr for f in findings):
            return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
